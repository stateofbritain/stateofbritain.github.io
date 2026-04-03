"""
parse-construction.py

Parses ONS construction output and new orders Excel files.

Usage:
  python3 scripts/parse-construction.py <output.xlsx> <orders.xlsx>

Outputs JSON to stdout.
"""

import sys
import json
import re
import openpyxl


def parse_construction_output(path):
    """
    Parse ONS bulletindataset2.xlsx, Table 2a.
    Annual rows (1997-2025) at rows 7-35, columns: year, then sector values.
    CDIDs at row 6 identify which column is which sector.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    ws = wb["Table 2a"]

    target_cdids = {
        "MV3W": "publicNewHousing",
        "MV3X": "privateNewHousing",
        "MV3Y": "infrastructure",
        "MV3Z": "publicOtherNewWork",
        "MV42": "privateIndustrial",
        "MV43": "privateCommercial",
        "MV44": "allNewWork",
    }

    # Row 6 has CDIDs, find column mapping
    col_map = {}
    rows = list(ws.iter_rows(min_row=1, max_row=40, values_only=False))

    # Find CDID row
    cdid_row_idx = None
    for i, row in enumerate(rows):
        for cell in row:
            val = str(cell.value or "").strip()
            if val in target_cdids:
                cdid_row_idx = i
                break
        if cdid_row_idx is not None:
            break

    if cdid_row_idx is None:
        print("Could not find CDID row", file=sys.stderr)
        wb.close()
        return []

    for cell in rows[cdid_row_idx]:
        val = str(cell.value or "").strip()
        if val in target_cdids:
            col_map[cell.column - 1] = target_cdids[val]  # 0-indexed

    print(f"Found {len(col_map)} columns at row {cdid_row_idx + 1}", file=sys.stderr)

    # Read annual data rows (pure 4-digit years)
    year_pattern = re.compile(r"^\d{4}$")
    annual = []

    for row in rows[cdid_row_idx + 1:]:
        date_val = str(row[0].value or "").strip()
        if not year_pattern.match(date_val):
            if annual:  # We've passed the annual section
                break
            continue

        entry = {"year": int(date_val)}
        for col_idx, key in col_map.items():
            val = row[col_idx].value
            if val is not None:
                try:
                    entry[key] = int(float(val))
                except (ValueError, TypeError):
                    entry[key] = None
            else:
                entry[key] = None
        annual.append(entry)

    wb.close()
    return annual


def parse_health_orders(path):
    """
    Parse ONS bulletindataset7.xlsx — Table 5 has Health sub-category.
    Quarterly columns like "Jan to Mar 1985", "Apr to Jun 1985".
    We aggregate to annual totals.
    """
    wb = openpyxl.load_workbook(path, data_only=True, read_only=True)

    target_sheet = None
    for name in wb.sheetnames:
        if name.strip() in ("5", "Table 5"):
            target_sheet = name
            break
    if not target_sheet:
        for name in wb.sheetnames:
            if "5" in name:
                target_sheet = name
                break

    if not target_sheet:
        print(f"Table 5 not found. Sheets: {wb.sheetnames}", file=sys.stderr)
        wb.close()
        return []

    print(f"Using orders sheet: {target_sheet}", file=sys.stderr)
    ws = wb[target_sheet]
    rows = list(ws.iter_rows(values_only=False))

    # Find Health row
    health_row = None
    for i, row in enumerate(rows):
        cell0 = str(row[0].value or "").strip().lower()
        if cell0 == "health":
            health_row = row
            print(f"  Found 'Health' at row {i+1}", file=sys.stderr)
            break

    # Find header row with quarterly dates like "Jan to Mar 1985"
    header_row = None
    quarter_pattern = re.compile(r"(Jan|Apr|Jul|Oct)\s+to\s+\w+\s+(\d{4})", re.I)
    for i, row in enumerate(rows):
        matches = sum(1 for c in row if quarter_pattern.search(str(c.value or "")))
        if matches >= 5:
            header_row = row
            print(f"  Found header at row {i+1} ({matches} quarterly columns)", file=sys.stderr)
            break

    if health_row is None or header_row is None:
        print("Could not find Health row or header", file=sys.stderr)
        wb.close()
        return []

    # Extract quarterly data and aggregate to annual
    yearly = {}
    for j, header_cell in enumerate(header_row):
        hval = str(header_cell.value or "")
        m = quarter_pattern.search(hval)
        if not m or j >= len(health_row):
            continue
        year = int(m.group(2))
        val = health_row[j].value
        if val is not None:
            try:
                fval = float(val)
                if year not in yearly:
                    yearly[year] = {"sum": 0, "count": 0}
                yearly[year]["sum"] += fval
                yearly[year]["count"] += 1
            except (ValueError, TypeError):
                pass

    # Only include years with all 4 quarters
    result = []
    for year in sorted(yearly.keys()):
        if yearly[year]["count"] == 4:
            result.append({"year": year, "healthOrders": int(yearly[year]["sum"])})

    wb.close()
    return result


def main():
    if len(sys.argv) < 3:
        print("Usage: python3 parse-construction.py <output.xlsx> <orders.xlsx>", file=sys.stderr)
        sys.exit(1)

    construction = parse_construction_output(sys.argv[1])
    print(f"Construction output: {len(construction)} annual rows", file=sys.stderr)
    if construction:
        print(f"  Range: {construction[0]['year']}-{construction[-1]['year']}", file=sys.stderr)
        print(f"  Latest: {construction[-1]}", file=sys.stderr)

    health_orders = parse_health_orders(sys.argv[2])
    print(f"Health orders: {len(health_orders)} rows", file=sys.stderr)
    if health_orders:
        print(f"  Range: {health_orders[0]['year']}-{health_orders[-1]['year']}", file=sys.stderr)
        print(f"  Latest: {health_orders[-1]}", file=sys.stderr)

    json.dump({
        "constructionOutput": construction,
        "healthOrders": health_orders,
    }, sys.stdout, indent=2)


if __name__ == "__main__":
    main()
